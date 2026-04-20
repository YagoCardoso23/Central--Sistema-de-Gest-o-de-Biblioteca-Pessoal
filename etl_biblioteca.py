# =============================================================================
#  ETL AUTOMÁTICO – SISTEMA DE BIBLIOTECA
#  Versão local — roda no VS Code / terminal Windows, Mac ou Linux
# =============================================================================
#
#  PASSO 1 — instale as dependências (só precisa fazer uma vez):
#    pip install pandas openpyxl
#
#  PASSO 2 — coloque este arquivo e as pastas assim:
#
#   sistemaetl\                  <- pasta raiz (pode ter qualquer nome)
#   ├── etl_biblioteca.py        <- este arquivo
#   ├── arquivoscsv\
#   │   ├── usuarios_100.csv
#   │   ├── livros.csv
#   │   └── planos.csv
#   └── arquivosxml\
#       ├── historico_mes_01.xml
#       └── ... (até o mês 12)
#
#  PASSO 3 — execute:
#    python etl_biblioteca.py
#
#  O relatório será salvo na mesma pasta do script.
# =============================================================================

import os
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Detecta automaticamente a pasta onde este script está salvo
DIRETORIO_BASE = Path(__file__).parent

PASTA_CSV     = DIRETORIO_BASE / 'arquivoscsv'
PASTA_XML     = DIRETORIO_BASE / 'arquivosxml'
ARQUIVO_SAIDA = DIRETORIO_BASE / 'relatorio_biblioteca_etl.xlsx'


# =============================================================================
#  EXTRACT
# =============================================================================

def carregar_csvs(pasta: Path):
    df_usuarios = pd.read_csv(pasta / 'usuarios_100.csv')
    df_livros   = pd.read_csv(pasta / 'livros.csv')
    df_planos   = pd.read_csv(pasta / 'planos.csv')
    print(f"  OK {len(df_usuarios)} usuarios | {len(df_livros)} livros | {len(df_planos)} planos")
    return df_usuarios, df_livros, df_planos


def carregar_xmls(pasta: Path) -> pd.DataFrame:
    if not pasta.exists():
        raise FileNotFoundError(f"Pasta nao encontrada: {pasta}")

    registros, arquivos_lidos = [], 0
    for arq in sorted(pasta.iterdir()):
        if arq.suffix.lower() != '.xml':
            continue
        tree = ET.parse(arq)
        root = tree.getroot()
        for emp in root:
            datas = emp.find('DATAS')
            registros.append({
                'ID_Aluguel':      emp.attrib.get('ID_Aluguel'),
                'LIVRO_ID':        emp.findtext('LIVRO_ID'),
                'USUARIO_LOGIN':   emp.findtext('USUARIO_LOGIN'),
                'DATA_EMPRESTIMO': datas.findtext('DATA_EMPRESTIMO') if datas is not None else None,
                'DATA_DEVOL_PREV': datas.findtext('DATA_DEVOLUCAO_PREVISTA') if datas is not None else None,
                'STATUS':          emp.findtext('STATUS'),
            })
        arquivos_lidos += 1

    print(f"  OK {arquivos_lidos} arquivos XML | {len(registros)} emprestimos extraidos")
    return pd.DataFrame(registros)


# =============================================================================
#  TRANSFORM
# =============================================================================

def transformar(df_emp, df_usuarios, df_livros, df_planos):
    hoje = pd.to_datetime(datetime.now().date())

    df_emp['LIVRO_ID']        = df_emp['LIVRO_ID'].astype(str)
    df_emp['DATA_EMPRESTIMO'] = pd.to_datetime(df_emp['DATA_EMPRESTIMO'])
    df_emp['DATA_DEVOL_PREV'] = pd.to_datetime(df_emp['DATA_DEVOL_PREV'])
    df_livros['ID_Livro']     = df_livros['ID_Livro'].astype(str)

    df = df_emp.merge(df_usuarios, left_on='USUARIO_LOGIN', right_on='Login', how='left')
    df = df.merge(df_livros,       left_on='LIVRO_ID',      right_on='ID_Livro', how='left')
    df = df.merge(df_planos,       left_on='Plano_ID',      right_on='ID_Plano', how='left')
    df = df.rename(columns={'Tipo_Plano_x': 'Tipo_Plano'}).drop(columns=['Tipo_Plano_y'], errors='ignore')

    df['dias_posse']    = (hoje - df['DATA_EMPRESTIMO']).dt.days
    df['dias_atraso']   = (hoje - df['DATA_DEVOL_PREV']).dt.days.clip(lower=0)
    df['atrasado']      = (df['STATUS'] == 'Ativo') & (hoje > df['DATA_DEVOL_PREV'])
    df['multa_R$']      = df.apply(
        lambda r: round(r['dias_atraso'] * r['Multa_Diaria'], 2) if r['atrasado'] else 0.0,
        axis=1
    )
    df['Status_Atraso'] = df['atrasado'].map({True: 'SIM', False: 'NAO'})

    print(f"  OK {len(df)} registros | {int(df['atrasado'].sum())} atrasados | Multa: R$ {df['multa_R$'].sum():,.2f}")
    return df, hoje


def montar_abas(df: pd.DataFrame) -> dict:
    princ = df[[
        'ID_Aluguel','USUARIO_LOGIN','Nome','Sobrenome','Tipo_Plano',
        'Titulo','Autor','Genero','DATA_EMPRESTIMO','DATA_DEVOL_PREV',
        'STATUS','dias_posse','dias_atraso','Status_Atraso','multa_R$'
    ] if 'Titulo' in df.columns else [
        'ID_Aluguel','USUARIO_LOGIN','Nome','Sobrenome','Tipo_Plano',
        'Título','Autor','Gênero','DATA_EMPRESTIMO','DATA_DEVOL_PREV',
        'STATUS','dias_posse','dias_atraso','Status_Atraso','multa_R$'
    ]].copy()
    princ.columns = [
        'ID Aluguel','Login','Nome','Sobrenome','Plano',
        'Titulo do Livro','Autor','Genero','Data Emprestimo','Data Prev. Devolucao',
        'Status','Dias em Posse','Dias de Atraso','Em Atraso?','Multa (R$)'
    ]
    princ['Data Emprestimo']      = princ['Data Emprestimo'].dt.strftime('%d/%m/%Y')
    princ['Data Prev. Devolucao'] = princ['Data Prev. Devolucao'].dt.strftime('%d/%m/%Y')

    titulo_col = 'Título' if 'Título' in df.columns else 'Titulo'
    atrasos = df[df['atrasado']][[
        'ID_Aluguel','Nome','Sobrenome','USUARIO_LOGIN','Tipo_Plano',
        titulo_col,'DATA_EMPRESTIMO','DATA_DEVOL_PREV','dias_atraso','multa_R$'
    ]].sort_values('multa_R$', ascending=False).copy()
    atrasos.columns = [
        'ID Aluguel','Nome','Sobrenome','Login','Plano',
        'Livro','Data Emprestimo','Data Prev. Devolucao','Dias de Atraso','Multa (R$)'
    ]
    atrasos['Data Emprestimo']      = atrasos['Data Emprestimo'].dt.strftime('%d/%m/%Y')
    atrasos['Data Prev. Devolucao'] = atrasos['Data Prev. Devolucao'].dt.strftime('%d/%m/%Y')

    genero_col = 'Gênero' if 'Gênero' in df.columns else 'Genero'
    livros = (df.groupby([titulo_col,'Autor', genero_col])
                .agg(Total=('ID_Aluguel','count'), Atrasos=('atrasado','sum'))
                .reset_index().sort_values('Total', ascending=False).head(30))
    livros.columns = ['Titulo','Autor','Genero','Total Emprestimos','Vezes com Atraso']

    planos = (df.groupby('Tipo_Plano')
                .agg(Total=('ID_Aluguel','count'),
                     Atrasados=('atrasado','sum'),
                     Multa_Total=('multa_R$','sum'),
                     Usuarios=('USUARIO_LOGIN','nunique'))
                .reset_index().sort_values('Multa_Total', ascending=False))
    planos.columns = ['Plano','Total Emprestimos','Qtd. Atrasados','Multa Total (R$)','Usuarios Unicos']

    usuarios = (df.groupby(['USUARIO_LOGIN','Nome','Sobrenome','Tipo_Plano','Status_Ativo'])
                  .agg(Total=('ID_Aluguel','count'),
                       Atrasos=('atrasado','sum'),
                       Multa=('multa_R$','sum'))
                  .reset_index().sort_values('Total', ascending=False).head(30))
    usuarios.columns = ['Login','Nome','Sobrenome','Plano','Status','Total Emprestimos','Atrasos','Multa Total (R$)']

    return {
        'Emprestimos':  princ,
        'Atrasos':      atrasos,
        'Top Livros':   livros,
        'Por Plano':    planos,
        'Top Usuarios': usuarios,
    }


# =============================================================================
#  LOAD
# =============================================================================

COR_HEADER = '1F4E79'
COR_AZ1    = 'DEEAF1'
COR_AZ2    = 'FFFFFF'
COR_TITULO = '2E75B6'


def _borda():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


def _formatar_header(ws):
    for cell in ws[1]:
        cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        cell.fill      = PatternFill('solid', start_color=COR_HEADER)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = _borda()
    ws.row_dimensions[1].height = 28


def _ajustar_colunas(ws):
    for col in ws.columns:
        w = max(10, min(45, max(len(str(c.value or '')) for c in col) + 4))
        ws.column_dimensions[get_column_letter(col[0].column)].width = w


def _escrever_aba(ws, df_aba: pd.DataFrame):
    for r in dataframe_to_rows(df_aba, index=False, header=True):
        ws.append(r)
    _formatar_header(ws)
    f1 = PatternFill('solid', start_color=COR_AZ1)
    f2 = PatternFill('solid', start_color=COR_AZ2)
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font      = Font(name='Arial', size=10)
            cell.border    = _borda()
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill      = f1 if i % 2 == 0 else f2
    _ajustar_colunas(ws)
    ws.freeze_panes    = 'A2'
    ws.auto_filter.ref = ws.dimensions


def _colorir_coluna_atraso(ws):
    col_idx = None
    for cell in ws[1]:
        if cell.value == 'Em Atraso?':
            col_idx = cell.column
            break
    if not col_idx:
        return
    fill_sim = PatternFill('solid', start_color='FFEB9C')
    fill_nao = PatternFill('solid', start_color='C6EFCE')
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        for cell in row:
            cell.fill = fill_sim if 'SIM' in str(cell.value or '') else fill_nao


def _criar_aba_resumo(wb, df: pd.DataFrame, hoje):
    ws = wb.create_sheet('Resumo', 0)
    ws.sheet_view.showGridLines = False

    linhas = [
        ['RELATORIO ETL - SISTEMA DE BIBLIOTECA', ''],
        [f'Gerado em: {hoje.strftime("%d/%m/%Y")}', ''],
        ['', ''],
        ['METRICA', 'VALOR'],
        ['Total de Registros no Historico',  len(df)],
        ['Emprestimos Ativos',               int((df['STATUS'] == 'Ativo').sum())],
        ['Emprestimos Devolvidos',           int((df['STATUS'] == 'Devolvido').sum())],
        ['Emprestimos em Atraso',            int(df['atrasado'].sum())],
        ['Multa Total Acumulada (R$)',        round(df['multa_R$'].sum(), 2)],
        ['Usuarios Unicos com Atraso',       int(df[df['atrasado']]['USUARIO_LOGIN'].nunique())],
        ['Livros Unicos Emprestados',        int(df['LIVRO_ID'].nunique())],
        ['Meses de Historico Processados',  12],
        ['Arquivos XML Carregados',          12],
    ]

    for i, linha in enumerate(linhas, 1):
        ws.append(linha)

    ws['A1'].font = Font(name='Arial', bold=True, size=16, color=COR_TITULO)
    ws['A2'].font = Font(name='Arial', italic=True, size=10, color='7F7F7F')
    ws.row_dimensions[1].height = 40

    for cell in ws[4]:
        cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=12)
        cell.fill      = PatternFill('solid', start_color=COR_HEADER)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = _borda()
        ws.row_dimensions[4].height = 28

    f_az = PatternFill('solid', start_color=COR_AZ1)
    f_br = PatternFill('solid', start_color=COR_AZ2)
    for i in range(5, len(linhas) + 1):
        for cell in ws[i]:
            cell.border    = _borda()
            cell.fill      = f_az if i % 2 == 0 else f_br
            cell.font      = Font(name='Arial', size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if cell.column == 2:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.font      = Font(name='Arial', bold=True, size=11)
        ws.row_dimensions[i].height = 22

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25


def gerar_excel(df: pd.DataFrame, abas: dict, hoje, caminho_saida: Path):
    wb = Workbook()
    wb.remove(wb.active)

    for nome, df_aba in abas.items():
        ws = wb.create_sheet(nome)
        _escrever_aba(ws, df_aba)

    _colorir_coluna_atraso(wb['Emprestimos'])
    _criar_aba_resumo(wb, df, hoje)

    wb.save(caminho_saida)
    print(f"  OK Arquivo salvo em: {caminho_saida}")


# =============================================================================
#  PIPELINE PRINCIPAL
# =============================================================================

if __name__ == '__main__':
    print("\nIniciando pipeline ETL - Biblioteca\n")
    try:
        print("[1/4] Extraindo CSVs...")
        df_usuarios, df_livros, df_planos = carregar_csvs(PASTA_CSV)

        print("\n[2/4] Extraindo XMLs...")
        df_emp = carregar_xmls(PASTA_XML)

        print("\n[3/4] Transformando dados...")
        df_final, hoje = transformar(df_emp, df_usuarios, df_livros, df_planos)
        abas = montar_abas(df_final)

        print("\n[4/4] Gerando Excel...")
        gerar_excel(df_final, abas, hoje, ARQUIVO_SAIDA)

        print(f"\nETL concluido! Abra o arquivo:\n   {ARQUIVO_SAIDA}\n")

    except FileNotFoundError as e:
        print(f"\nERRO - Arquivo ou pasta nao encontrado: {e}")
        print(f"\n  Verifique se a estrutura esta assim:")
        print(f"  {DIRETORIO_BASE}")
        print(f"  arquivoscsv\\  (com os 3 CSVs)")
        print(f"  arquivosxml\\ (com os 12 XMLs)\n")
    except ModuleNotFoundError as e:
        print(f"\nERRO - Biblioteca nao instalada: {e}")
        print(f"\n  Rode no terminal:\n  pip install pandas openpyxl\n")
    except Exception as e:
        print(f"\nERRO inesperado: {e}")
        raise
